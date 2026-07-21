#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Global FCFP AI Productivity — Use Case Hub · local backend
===========================================================

Why this exists
---------------
A browser (SheetJS) can only REBUILD an .xlsx when it saves, which strips
conditional formatting, data-validation dropdowns, column widths, merged
cells and cell styles. This backend instead opens the REAL file with
openpyxl and only writes the new row, so all existing formatting and UI
elements in the tracker are preserved. New rows also inherit the style of
the row above them, and dropdown validations are extended to cover them.

How to run
----------
1. Install Python 3.9+  (on Windows: from python.org or the Store)
2. In a terminal / PowerShell, in this folder:
       pip install -r requirements.txt
3. Point TRACKER_PATH below at your Excel file. If that file lives in a
   synced SharePoint / OneDrive folder, edits sync back automatically.
4. Start it:
       python server.py
5. Open the address it prints (default http://127.0.0.1:5000) in Edge/Chrome.

The web page auto-detects the backend; no manual file loading needed.
If you open the .html on its own without this server, it still works in
the old download-a-copy mode.
"""

import os
import re
import datetime
import threading
import webbrowser
from copy import copy

from flask import Flask, request, jsonify, send_from_directory
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# ─────────────────────────── CONFIG ───────────────────────────
# Absolute or relative path to the tracker. Use a raw string on Windows,
# e.g. r"C:\Users\you\ING\Global FCFP\UseCaseTracker.xlsx"
TRACKER_PATH = os.environ.get("TRACKER_PATH", "UseCaseTracker.xlsx")

# Preferred sheet/tab name. "" = auto-detect the sheet that looks like the
# use-case table.
SHEET_NAME = os.environ.get("SHEET_NAME", "")

# The HTML file served to the browser (kept next to this script).
HTML_FILE = os.environ.get("HTML_FILE", "fcfp-ai-usecase-hub.html")

HOST = os.environ.get("HOST", "127.0.0.1")
PORT = int(os.environ.get("PORT", "5000"))
# ───────────────────────────────────────────────────────────────

# Canonical fields — MUST stay in sync with the FIELDS array in the HTML.
FIELDS = [
    ("num",            "Use Case #", 0, ["use case #","use case no","use case number","usecase number","use case id","uc #","uc no","uc id","case number","case no","case id","sr no","sr","s no","sl no","sl","serial","serial no","index","row","no","nr","id","number"]),
    ("name",           "Use Case Name", 0, ["use case name","use case title","usecase name","name","title","use case"]),
    ("domain",         "Domain", 0, ["domain","area","business area","function"]),
    ("sponsor",        "Sponsor", 0, ["sponsor","project owner","business sponsor","project sponsor"]),
    ("aimaker",        "AI Maker", 0, ["ai maker","maker","ai developer","developer","builder","ai builder"]),
    ("translator",     "Business Translator", 0, ["business translator","translator"]),
    ("contact",        "Main Contact Person", 0, ["main contact person","main contact","contact person","point of contact","poc","contact"]),
    ("agentowner",     "Agent Owner", 0, ["agent owner","owner of agent","accountable owner","accountability"]),
    ("problem",        "Problem Statement", 1, ["problem statement","problem","pain point","challenge","issue"]),
    ("current",        "Current Process", 1, ["current process","as is process","as-is","current state","existing process","current way"]),
    ("agentdo",        "What Agent Will Do", 1, ["what agent will do","what the agent will do","agent will do","agent scope","proposed solution","to be process","solution"]),
    ("datasources",    "Data Sources & Input Data Type", 1, ["data sources & input data type","data sources and input","data sources","data source","input data type","input data","data type"]),
    ("classification", "Data Classification", 1, ["data classification","classification","data class","confidentiality"]),
    ("driver",         "Driver / Urgency", 1, ["driver / urgency","driver urgency","driver","urgency","reason"]),
    ("value",          "Estimated Value", 1, ["estimated value","expected value","value estimate","benefit","impact","business value"]),
    ("targets",        "Target Users / Teams", 1, ["target users / teams","target users","target teams","target user","target team","beneficiaries"]),
    ("similar",        "Existing Similar Builds", 1, ["existing similar builds","existing similar","similar builds","similar build","similar use case","duplicate"]),
    ("zone",           "Copilot Zone", 1, ["copilot zone","co-pilot zone","zone"]),
    ("zonejust",       "Zone Justification", 1, ["zone justification","zone reason","justification for zone","justification"]),
    ("agentmgmt",      "Main Responsible for Agent Management", 1, ["main responsible for agent management","responsible for agent management","agent management","maintenance owner","who maintains"]),
    ("poapproval",     "Process Owner Approval", 2, ["process owner approval","po approval","process approval"]),
    ("rtarch",         "RT Architecture Review", 2, ["rt architecture review","architecture review","rt arch","architecture"]),
    ("pii",            "Personal Data / PII Involved", 2, ["personal data / pii involved","personal data pii","pii involved","personal data","pii"]),
    ("dtia",           "DTIA Approval", 2, ["dtia approval","dtia"]),
    ("euai",           "EU AI Act Assessment", 2, ["eu ai act assessment","eu ai act","ai act","euai"]),
    ("testing",        "Agent Testing", 2, ["agent testing","testing","tested","test status"]),
    ("demomt",         "Target Demo to MT", 3, ["target demo to mt","demo to mt","demo to management team","demo to management","target demo","management demo"]),
    ("iter1",          "Iteration 1", 3, ["iteration 1","iter 1","iteration1"]),
    ("demo1",          "Demo 1", 3, ["demo 1","demo1"]),
    ("iter2",          "Iteration 2", 3, ["iteration 2","iter 2","iteration2"]),
    ("demo2",          "Demo 2", 3, ["demo 2","demo2"]),
    ("rollout",        "Target Rollout", 3, ["target rollout","rollout","go live","go-live","deployment date","production date"]),
    ("tstatus",        "Timeline Status", 3, ["timeline status","timeline rag","rag status","track status"]),
    ("prioValue",      "Value", 4, ["value (priority)","priority value","value score"]),
    ("prioReuse",      "Reusability", 4, ["reusability","reuse","reusable"]),
    ("prioComplex",    "Complexity", 4, ["complexity","complex","effort"]),
    ("status",         "Status", 4, ["status","overall status","stage","state"]),
    ("notes",          "Notes / Comments", 4, ["notes / comments","notes","comments","remarks","note"]),
]
FIELD_IDS = [f[0] for f in FIELDS]
FIELD_LABEL = {f[0]: f[1] for f in FIELDS}

_norm_re = re.compile(r"[\s_\-/().,#:&]+")
def norm(s):
    if s is None:
        return ""
    return _norm_re.sub(" ", str(s).lower()).strip()

def to_str(val):
    """Render a cell value for the dashboard/JSON."""
    if val is None:
        return ""
    if isinstance(val, datetime.datetime):
        if val.hour == 0 and val.minute == 0 and val.second == 0:
            return val.strftime("%Y-%m-%d")
        return val.strftime("%Y-%m-%d %H:%M")
    if isinstance(val, datetime.date):
        return val.strftime("%Y-%m-%d")
    if isinstance(val, float) and val.is_integer():
        return str(int(val))
    return str(val)


# ─────────────── detection (mirrors the HTML JS) ───────────────
def detect_header(aoa):
    scan = min(len(aoa), 25)
    best_i, best_s = 0, -1
    for r in range(scan):
        cells = [norm(c) for c in aoa[r]]
        filled = [c for c in cells if c]
        if len(filled) < 2:
            continue
        s = len(filled) * 0.4
        seen = set()
        for fid, label, _grp, syn in FIELDS:
            if fid in seen:
                continue
            nlabel = norm(label)
            for c in filled:
                if c == nlabel:
                    s += 4; seen.add(fid); break
                if any(c == norm(a) for a in syn):
                    s += 3; seen.add(fid); break
                if any(len(norm(a)) >= 4 and norm(a) in c for a in syn):
                    s += 2; seen.add(fid); break
        s -= sum(1 for c in filled if len(c) > 60) * 1.5
        if s > best_s:
            best_s, best_i = s, r
    return best_i, best_s


def build_map(file_headers):
    H = [norm(h) for h in file_headers]
    cands = []
    for fid, label, _grp, syn in FIELDS:
        canon = norm(label)
        for ci, h in enumerate(H):
            if not h:
                continue
            sc = 0
            if h == canon:
                sc = 100
            else:
                for a in syn:
                    na = norm(a)
                    if not na:
                        continue
                    if h == na:
                        sc = max(sc, 92)
                    elif len(na) >= 4 and na in h:
                        sc = max(sc, 60 + len(na))
                    elif len(h) >= 4 and h in na:
                        sc = max(sc, 55 + len(h))
            if sc > 0:
                cands.append((sc, fid, ci))
    cands.sort(key=lambda x: -x[0])
    m, usedF, usedC = {}, set(), set()
    for sc, fid, ci in cands:
        if sc < 58 or fid in usedF or ci in usedC:
            continue
        m[fid] = ci; usedF.add(fid); usedC.add(ci)
    unmapped = [ci for ci in range(len(file_headers))
                if ci not in usedC and str(file_headers[ci]).strip() != ""]
    return m, unmapped


def ensure_number_column(m, file_headers, data_rows):
    if m.get("num") is not None:
        return
    used = set(m.values())
    best_ci, best_sc = -1, 0.0
    for ci in range(len(file_headers)):
        if ci in used:
            continue
        tot = ints = 0
        for r in data_rows:
            val = to_str(r[ci]).strip() if ci < len(r) else ""
            if not val:
                continue
            tot += 1
            if re.fullmatch(r"\d+", val):
                ints += 1
        if tot > 0:
            sc = ints / tot
            if sc >= 0.6 and sc > best_sc:
                best_sc, best_ci = sc, ci
    if best_ci > -1:
        m["num"] = best_ci
    elif 0 not in used:
        m["num"] = 0


def is_real_row(m, r):
    def g(fid):
        ci = m.get(fid)
        return to_str(r[ci]).strip() if (ci is not None and ci < len(r)) else ""
    if g("num") or g("name"):
        return True
    return sum(1 for c in r if to_str(c).strip() != "") >= 3


# ─────────────── workbook reading / writing ───────────────
def pick_sheet(wb):
    if SHEET_NAME:
        for n in wb.sheetnames:
            if norm(n) == norm(SHEET_NAME):
                return n
    best, best_s = wb.sheetnames[0], -1
    for n in wb.sheetnames:
        ws = wb[n]
        aoa = [list(row) for row in ws.iter_rows(values_only=True)]
        _, s = detect_header(aoa)
        if s > best_s:
            best_s, best = s, n
    return best


def read_tracker(sheet_override=None):
    """Load the file and return the same structures the HTML builds."""
    if not os.path.exists(TRACKER_PATH):
        return {"error": "TRACKER_PATH not found: %s" % os.path.abspath(TRACKER_PATH)}
    wb = load_workbook(TRACKER_PATH, data_only=True, read_only=True)
    try:
        sheet = sheet_override if (sheet_override in wb.sheetnames) else pick_sheet(wb)
        ws = wb[sheet]
        aoa = [list(row) for row in ws.iter_rows(values_only=True)]
    finally:
        wb.close()

    hidx, _ = detect_header(aoa)
    raw_headers = aoa[hidx] if hidx < len(aoa) else []
    file_headers = [to_str(h).strip() for h in raw_headers]

    data_rows = []
    for row in aoa[hidx + 1:]:
        r = [to_str(c) for c in row]
        while len(r) < len(file_headers):
            r.append("")
        data_rows.append(r)

    m, unmapped = build_map(file_headers)
    ensure_number_column(m, file_headers, data_rows)
    rows = [r for r in data_rows if is_real_row(m, r)]

    # next use case number
    mx = 0
    ci = m.get("num")
    if ci is not None:
        for r in rows:
            v = r[ci] if ci < len(r) else ""
            if re.fullmatch(r"\d+", str(v).strip() or ""):
                mx = max(mx, int(v))
    next_number = str(mx + 1)

    return {
        "fileName": os.path.basename(TRACKER_PATH),
        "filePath": os.path.abspath(TRACKER_PATH),
        "sheetName": sheet,
        "sheetNames": wb.sheetnames if hasattr(wb, "sheetnames") else [sheet],
        "headerRowIndex": hidx,
        "headers": file_headers,
        "rows": rows,
        "map": m,
        "unmappedCols": unmapped,
        "nextNumber": next_number,
    }


# openpyxl read_only workbooks don't expose sheetnames after close; re-open lightly
def _sheet_names():
    wb = load_workbook(TRACKER_PATH, read_only=True)
    try:
        return list(wb.sheetnames)
    finally:
        wb.close()


def coerce_value(fid, text):
    """Turn form text into a typed value where useful (dates, integers)."""
    if text is None:
        return None
    text = str(text).strip()
    if text == "":
        return None
    if fid == "num" and re.fullmatch(r"\d+", text):
        return int(text)
    # date fields come from <input type=date> as YYYY-MM-DD
    if fid in ("demomt", "iter1", "demo1", "iter2", "demo2", "rollout"):
        m = re.fullmatch(r"(\d{4})-(\d{2})-(\d{2})", text)
        if m:
            try:
                return datetime.date(int(m[1]), int(m[2]), int(m[3]))
            except ValueError:
                pass
    return text


def append_usecase(payload):
    """Append one use case to the REAL file, preserving formatting."""
    if not os.path.exists(TRACKER_PATH):
        return {"ok": False, "error": "TRACKER_PATH not found: %s" % os.path.abspath(TRACKER_PATH)}

    # figure out sheet + layout from a light read first
    meta = read_tracker(payload.get("_sheet"))
    if "error" in meta:
        return {"ok": False, "error": meta["error"]}

    sheet = meta["sheetName"]
    hidx = meta["headerRowIndex"]          # 0-based
    header_excel_row = hidx + 1            # 1-based
    file_headers = list(meta["headers"])
    m = dict(meta["map"])

    # assign next number if the user left it blank
    num_text = str(payload.get("num", "")).strip()
    if not num_text:
        num_text = meta["nextNumber"]
    payload = dict(payload)
    payload["num"] = num_text

    try:
        wb = load_workbook(TRACKER_PATH)  # full load — keeps styles/formatting
    except Exception as e:
        return {"ok": False, "error": "Could not open the file: %s" % e}

    if sheet not in wb.sheetnames:
        sheet = wb.sheetnames[0]
    ws = wb[sheet]

    # any populated field the file lacks → add as a new column at the header row
    for fid, label, _grp, _syn in FIELDS:
        if m.get(fid) is None and str(payload.get(fid, "")).strip() != "":
            new_ci = len(file_headers)
            file_headers.append(label)
            ws.cell(row=header_excel_row, column=new_ci + 1, value=label)
            m[fid] = new_ci

    # find the last real data row (by identity columns), append below it
    num_ci = m.get("num")
    name_ci = m.get("name")
    last_row = header_excel_row
    for r in range(header_excel_row + 1, ws.max_row + 1):
        got = False
        for ci in (num_ci, name_ci):
            if ci is not None:
                val = ws.cell(row=r, column=ci + 1).value
                if val is not None and str(val).strip() != "":
                    got = True
                    break
        if got:
            last_row = r
    new_row = last_row + 1

    # write values, inheriting the style of the row above so borders / fills /
    # number formats match the rest of the table
    style_src_row = last_row if last_row > header_excel_row else None
    for fid, label, _grp, _syn in FIELDS:
        ci = m.get(fid)
        if ci is None:
            continue
        val = coerce_value(fid, payload.get(fid, ""))
        cell = ws.cell(row=new_row, column=ci + 1, value=val)
        if style_src_row is not None:
            src = ws.cell(row=style_src_row, column=ci + 1)
            if src.has_style:
                cell.font = copy(src.font)
                cell.fill = copy(src.fill)
                cell.border = copy(src.border)
                cell.alignment = copy(src.alignment)
                cell.number_format = src.number_format
                cell.protection = copy(src.protection)

    # extend data-validation dropdowns (and other DVs) to cover the new row
    try:
        for dv in list(ws.data_validations.dataValidation):
            add_ranges = []
            for rng in list(dv.sqref.ranges):
                if rng.max_row == last_row and rng.min_row <= last_row:
                    for col in range(rng.min_col, rng.max_col + 1):
                        add_ranges.append("%s%d" % (get_column_letter(col), new_row))
            for a in add_ranges:
                dv.add(a)
    except Exception:
        pass  # validation extension is best-effort; never block the save

    try:
        wb.save(TRACKER_PATH)
    except PermissionError:
        return {"ok": False,
                "error": "The file is open in Excel (or locked). Close it and try again — this keeps your formatting intact."}
    except Exception as e:
        return {"ok": False, "error": "Save failed: %s" % e}
    finally:
        wb.close()

    tracker = read_tracker(sheet)
    return {"ok": True, "number": num_text, "fileName": meta["fileName"], "tracker": tracker}


# ─────────────────────────── Flask app ───────────────────────────
app = Flask(__name__)


@app.after_request
def no_cache(resp):
    resp.headers["Cache-Control"] = "no-store"
    return resp


@app.route("/")
def index():
    return send_from_directory(os.path.dirname(os.path.abspath(HTML_FILE)) or ".",
                               os.path.basename(HTML_FILE))


@app.route("/api/tracker")
def api_tracker():
    sheet = request.args.get("sheet")
    data = read_tracker(sheet)
    # attach full sheet list (read_tracker closes the read-only wb early)
    try:
        data["sheetNames"] = _sheet_names()
    except Exception:
        pass
    return jsonify(data)


@app.route("/api/usecase", methods=["POST"])
def api_usecase():
    payload = request.get_json(force=True, silent=True) or {}
    return jsonify(append_usecase(payload))


if __name__ == "__main__":
    print("=" * 64)
    print(" Global FCFP AI Productivity — Use Case Hub (local backend)")
    print("=" * 64)
    print(" Tracker : %s" % os.path.abspath(TRACKER_PATH))
    print(" Sheet   : %s" % (SHEET_NAME or "(auto-detect)"))
    print(" HTML    : %s" % os.path.abspath(HTML_FILE))
    print(" Open    : http://%s:%d" % (HOST, PORT))
    print("=" * 64)
    if not os.path.exists(TRACKER_PATH):
        print(" WARNING: tracker file not found yet — set TRACKER_PATH.")
    
    # Auto-open localhost URL in default browser 1.2s after server starts
    def open_browser():
        try:
            webbrowser.open(f"http://{HOST}:{PORT}")
        except Exception:
            pass

    threading.Timer(1.2, open_browser).start()
    app.run(host=HOST, port=PORT, debug=False)
